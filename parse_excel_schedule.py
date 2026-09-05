#!/usr/bin/env python3
"""
Скрипт імпорту розкладу занять з Excel ("Розклад І семестр 26-27 н.р. на 01.09.26.xlsx")
у формат schedule.json для Telegram-бота та сайту mpmek.site, а також оновлення OOBJECT.py.
"""

import json
import os
import re
import urllib.request
from pathlib import Path
from typing import Any, Dict, List, Set, Tuple

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
EXCEL_FILE = PROJECT_ROOT / "Розклад І семестр 26-27 н.р. на 01.09.26.xlsx"
SCHEDULE_OUT_BOT = PROJECT_ROOT / "schedule.json"
SCHEDULE_OUT_WEB = Path("/Users/nazariyshikircloud.com/Какая то идея/app/schedule.json")
OOBJECT_FILE = PROJECT_ROOT / "OOBJECT.py"
REMOTE_SCHEDULE_URL = "https://mpmek.site/schedule.json"

PAIR_MAP = {
    '0': 0, 'I': 1, '1': 1, 'II': 2, '2': 2,
    'III': 3, '3': 3, 'IV': 4, '4': 4, 'V': 5, 'VI': 6,
}

DAY_MAP = {
    'понеділок': 'Понеділок',
    'вівторок': 'Вівторок',
    'середа': 'Середа',
    'четвер': 'Четвер',
    "п'ятниця": "П'ятниця",
    'пʼятниця': "П'ятниця",
    'пятниця': "П'ятниця",
}

DAYS_ORDER = ['Понеділок', 'Вівторок', 'Середа', 'Четвер', "П'ятниця"]


def clean_subject(name: Any) -> str:
    s = str(name or '').strip()
    s = re.sub(r'\s+', ' ', s)
    if s.startswith('(') and s.count('(') > s.count(')'):
        s = s[1:].strip()
    if s == 'Біологіяі екологія':
        s = 'Біологія і екологія'
    elif s == 'ІсторіяУкраїни':
        s = 'Історія України'
    return s


def clean_teacher(name: Any) -> str:
    s = str(name or '').strip()
    return re.sub(r'\s+', ' ', s)


def clean_room(room: Any) -> str:
    if room is None:
        return ''
    return str(room).strip()


def normalize_group_name(raw: str) -> str:
    g = re.sub(r'\s+', ' ', raw).strip()
    if g == 'БС -2026':
        return 'БС-2026'
    return g


def fetch_existing_schedule() -> Dict[str, Any]:
    """Завантажує поточний schedule.json з сайту або з локального веб-репо для збереження підвісок."""
    try:
        req = urllib.request.Request(
            REMOTE_SCHEDULE_URL,
            headers={"User-Agent": "MPMEK-Importer/1.0"}
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode('utf-8'))
            print(f"✅ Отримано існуючий розклад з {REMOTE_SCHEDULE_URL} ({len(data)} ключів)")
            return data
    except Exception as exc:
        print(f"⚠️ Не вдалося отримати розклад з URL ({exc}). Спроба читання з локального файлу...")

    if SCHEDULE_OUT_WEB.exists():
        try:
            with open(SCHEDULE_OUT_WEB, 'r', encoding='utf-8') as f:
                data = json.load(f)
                print(f"✅ Отримано існуючий розклад з {SCHEDULE_OUT_WEB}")
                return data
        except Exception as exc:
            print(f"⚠️ Помилка читання {SCHEDULE_OUT_WEB}: {exc}")

    return {}


def parse_schedule_from_excel(excel_path: Path) -> Tuple[Dict[str, Any], Dict[str, Set[str]]]:
    print(f"📖 Завантаження книги Excel: {excel_path.name}...")
    wb = openpyxl.load_workbook(str(excel_path), data_only=False)

    schedule_data: Dict[str, Any] = {}
    group_subjects: Dict[str, Set[str]] = {}

    course_sheets = [s for s in wb.sheetnames if 'курс' in s.lower()]
    sheets_to_process = course_sheets if course_sheets else wb.sheetnames

    for sname in sheets_to_process:
        ws = wb[sname]
        merged_ranges = ws.merged_cells.ranges

        # 1. Знаходимо групи у рядку 3
        groups_in_sheet: List[Tuple[int, str]] = []
        for col in range(3, ws.max_column + 1, 3):
            val = ws.cell(3, col).value
            if val and str(val).strip() and not str(val).strip().startswith('Заступник'):
                gname = normalize_group_name(str(val))
                groups_in_sheet.append((col, gname))
                if gname not in schedule_data:
                    schedule_data[gname] = {
                        'ЧИСЕЛЬНИК': {d: [] for d in DAYS_ORDER},
                        'ЗНАМЕННИК': {d: [] for d in DAYS_ORDER},
                        'ПІДВІСКА': []
                    }
                group_subjects.setdefault(gname, set())

        print(f" Аркуш '{sname}': знайдено {len(groups_in_sheet)} груп: {[g for _, g in groups_in_sheet]}")

        cur_day = None
        r = 4
        while r <= ws.max_row:
            dval = ws.cell(r, 1).value
            if dval and str(dval).strip().lower() in DAY_MAP:
                cur_day = DAY_MAP[str(dval).strip().lower()]

            pval = ws.cell(r, 2).value
            pclean = str(pval).strip() if pval is not None else ''
            pnum = PAIR_MAP.get(pclean)

            if pnum is not None and cur_day:
                r_num = r
                r_den = r + 1

                for col, gname in groups_in_sheet:
                    # Перевіряємо вертикальне злиття рядків r_num та r_den для даного стовпця
                    is_merged = any(
                        r_num in range(m.min_row, m.max_row + 1) and
                        col in range(m.min_col, m.max_col + 1) and
                        m.min_row != m.max_row
                        for m in merged_ranges
                    )

                    s_num = clean_subject(ws.cell(r_num, col).value)
                    t_num = clean_teacher(ws.cell(r_num, col + 1).value)
                    rm_num = clean_room(ws.cell(r_num, col + 2).value)

                    s_den = clean_subject(ws.cell(r_den, col).value)
                    t_den = clean_teacher(ws.cell(r_den, col + 1).value)
                    rm_den = clean_room(ws.cell(r_den, col + 2).value)

                    if is_merged and s_num and s_num != '—':
                        entry = {
                            'number': pnum,
                            'subject': s_num,
                            'teacher': t_num,
                            'room': rm_num
                        }
                        schedule_data[gname]['ЧИСЕЛЬНИК'][cur_day].append(entry)
                        schedule_data[gname]['ЗНАМЕННИК'][cur_day].append(entry)
                        group_subjects[gname].add(s_num)
                    else:
                        if s_num and s_num != '—':
                            schedule_data[gname]['ЧИСЕЛЬНИК'][cur_day].append({
                                'number': pnum,
                                'subject': s_num,
                                'teacher': t_num,
                                'room': rm_num
                            })
                            group_subjects[gname].add(s_num)
                        if s_den and s_den != '—':
                            schedule_data[gname]['ЗНАМЕННИК'][cur_day].append({
                                'number': pnum,
                                'subject': s_den,
                                'teacher': t_den,
                                'room': rm_den
                            })
                            group_subjects[gname].add(s_den)
                r += 2
            else:
                r += 1

    return schedule_data, group_subjects


def merge_pidvesky(
    new_schedule: Dict[str, Any],
    existing_schedule: Dict[str, Any]
) -> None:
    """Переносить та зберігає підвіски для всіх груп."""
    inherited_count = 0

    # Маппінг спадкування підвісок для реорганізованих/розділених груп
    split_group_mappings = {
        'БО-2025': 'БО (Ф)-2025',
        'Ф-2025': 'БО (Ф)-2025',
        'БС-2025-1': 'БС-2025',
        'БС-2025-2': 'БС-2025',
    }

    for gname, gdata in new_schedule.items():
        if gname.startswith('_'):
            continue

        existing_subs = []
        if gname in existing_schedule:
            existing_subs = existing_schedule[gname].get('ПІДВІСКА', [])
        elif gname in split_group_mappings:
            parent_g = split_group_mappings[gname]
            existing_subs = existing_schedule.get(parent_g, {}).get('ПІДВІСКА', [])
            if existing_subs:
                print(f"  ↪️  Група '{gname}' успадкувала {len(existing_subs)} підвісок від '{parent_g}'")

        if existing_subs:
            gdata['ПІДВІСКА'] = existing_subs
            inherited_count += len(existing_subs)
        else:
            gdata['ПІДВІСКА'] = []

    print(f"📌 Успішно перенесено та збережено {inherited_count} підвісок!")


def add_backward_compat_aliases(
    schedule_data: Dict[str, Any],
    group_subjects: Dict[str, Set[str]]
) -> None:
    """Додає аліаси для старих назв груп, щоб у поточних користувачів не ламався розклад."""
    aliases = {
        'БО (Ф)-2025': 'БО-2025',
        'БС-2025': 'БС-2025-1',
    }
    for old_name, new_name in aliases.items():
        if new_name in schedule_data and old_name not in schedule_data:
            schedule_data[old_name] = schedule_data[new_name]
            group_subjects[old_name] = group_subjects[new_name]
            print(f"  🔗 Створено аліас сумісності: '{old_name}' ➔ '{new_name}'")


def update_oobject_py(group_subjects: Dict[str, Set[str]]) -> None:
    """Генерує та оновлює OOBJECT.py з усіма групами та актуальними предметами."""
    print(f"📝 Оновлення {OOBJECT_FILE.name}...")

    # Читаємо існуючі предмети з OOBJECT.py для збереження можливих рідкісних дисциплін
    existing_objects: Dict[str, List[str]] = {}
    if OOBJECT_FILE.exists():
        try:
            import importlib.util
            spec = importlib.util.spec_from_file_location("old_oobject", str(OOBJECT_FILE))
            if spec and spec.loader:
                old_mod = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(old_mod)
                existing_objects = getattr(old_mod, "objects", {})
        except Exception as exc:
            print(f"⚠️ Не вдалося прочитати попередній OOBJECT.py: {exc}")

    merged_objects: Dict[str, Tuple[str, ...]] = {}

    # Спочатку додаємо всі нові групи
    for gname in sorted(group_subjects.keys()):
        cur_subs = set(group_subjects[gname])
        # Якщо група була раніше, об'єднуємо
        if gname in existing_objects:
            cur_subs.update(existing_objects[gname])
        merged_objects[gname] = tuple(sorted(cur_subs))

    # Зберігаємо старі групи яких немає в новому розкладі (наприклад, випускні БО-2023, Ф-2023)
    for gname, subs in existing_objects.items():
        if gname not in merged_objects:
            merged_objects[gname] = tuple(sorted(subs))

    # Форматуємо вміст OOBJECT.py
    lines = ["# -*- coding: utf-8 -*-", "objects = {"]
    for gname, subs in sorted(merged_objects.items()):
        lines.append(f'    "{gname}": (')
        for s in subs:
            lines.append(f'        "{s}",')
        lines.append('    ),')
        lines.append('')
    lines.append('}')
    lines.append('')

    content = '\n'.join(lines)
    with open(OOBJECT_FILE, 'w', encoding='utf-8') as f:
        f.write(content)
    print(f"✅ {OOBJECT_FILE.name} успішно оновлено! Загалом {len(merged_objects)} груп.")


def main():
    print("=" * 60)
    print("🚀 ІМПОРТ РОЗКЛАДУ В БОТА (mpmek)")
    print("=" * 60)

    if not EXCEL_FILE.exists():
        print(f"❌ Помилка: файл '{EXCEL_FILE}' не знайдено!")
        return

    # 1. Завантажуємо поточні підвіски
    existing_schedule = fetch_existing_schedule()

    # 2. Парсимо розклад з Excel
    new_schedule, group_subjects = parse_schedule_from_excel(EXCEL_FILE)
    print(f"📊 Успішно розпарсено {len(new_schedule)} груп з Excel!")

    # 3. Враховуємо та переносимо підвіски
    merge_pidvesky(new_schedule, existing_schedule)

    # 4. Зберігаємо налаштування сайту (_settings)
    if '_settings' in existing_schedule:
        new_schedule['_settings'] = existing_schedule['_settings']
    else:
        new_schedule['_settings'] = {
            "lessonTimes": {
                "0": "07:00 - 08:20",
                "1": "08:30 - 09:50",
                "2": "10:00 - 11:20",
                "3": "11:50 - 13:10",
                "4": "13:20 - 14:40",
                "5": "16:00 - 17:20",
                "6": "17:30 - 18:50"
            }
        }

    # 5. Додаємо аліаси для зворотної сумісності зі старими назвами
    add_backward_compat_aliases(new_schedule, group_subjects)

    # 6. Записуємо schedule.json для бота
    with open(SCHEDULE_OUT_BOT, 'w', encoding='utf-8') as f:
        json.dump(new_schedule, f, ensure_ascii=False, indent=2)
    print(f"💾 Розклад збережено у: {SCHEDULE_OUT_BOT} ({SCHEDULE_OUT_BOT.stat().st_size} байт)")

    # 7. Записуємо schedule.json у веб-репо, якщо доступне
    if SCHEDULE_OUT_WEB.parent.exists():
        with open(SCHEDULE_OUT_WEB, 'w', encoding='utf-8') as f:
            json.dump(new_schedule, f, ensure_ascii=False, indent=2)
        print(f"💾 Розклад збережено у веб-репо: {SCHEDULE_OUT_WEB}")

    # 8. Оновлюємо OOBJECT.py
    update_oobject_py(group_subjects)

    print("=" * 60)
    print("🎉 ІМПОРТ УСПІШНО ЗАВЕРШЕНО!")
    print("=" * 60)


if __name__ == "__main__":
    main()
